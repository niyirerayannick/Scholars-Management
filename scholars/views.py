from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from accounts.permissions import DATA_OFFICER, PROGRAM_MANAGER, can_edit_scholars, role_required

from .forms import ScholarForm, filter_scholars
from .models import Scholar


@login_required
def scholar_list(request):
    scholars, form = filter_scholars(Scholar.objects.all(), request.GET)
    if request.GET.get("export") == "excel":
        return export_scholars_excel(scholars)
    paginator = Paginator(scholars, 25)
    page_obj = paginator.get_page(request.GET.get("page"))
    query_params = request.GET.copy()
    query_params.pop("page", None)
    query_string = query_params.urlencode()
    return render(request, "scholars/list.html", {"page_obj": page_obj, "filter_form": form, "can_edit": can_edit_scholars(request.user), "query_string": query_string})


@login_required
def scholar_detail(request, pk):
    scholar = get_object_or_404(Scholar, pk=pk)
    return render(request, "scholars/detail.html", {"scholar": scholar, "can_edit": can_edit_scholars(request.user)})


@role_required(PROGRAM_MANAGER, DATA_OFFICER)
def scholar_create(request):
    form = ScholarForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        scholar = form.save()
        messages.success(request, "Scholar created successfully.")
        return redirect(scholar)
    return render(request, "scholars/form.html", {"form": form, "title": "Add Scholar"})


@role_required(PROGRAM_MANAGER, DATA_OFFICER)
def scholar_update(request, pk):
    scholar = get_object_or_404(Scholar, pk=pk)
    form = ScholarForm(request.POST or None, instance=scholar)
    if request.method == "POST" and form.is_valid():
        scholar = form.save()
        messages.success(request, "Scholar updated successfully.")
        return redirect(scholar)
    return render(request, "scholars/form.html", {"form": form, "title": "Edit Scholar"})


@role_required(PROGRAM_MANAGER)
def scholar_delete(request, pk):
    scholar = get_object_or_404(Scholar, pk=pk)
    if request.method == "POST":
        scholar.delete()
        messages.success(request, "Scholar deleted successfully.")
        return redirect("scholar_list")
    return render(request, "scholars/confirm_delete.html", {"scholar": scholar})


def export_scholars_excel(scholars):
    wb = Workbook()
    ws = wb.active
    ws.title = "Scholars"
    headers = ["Student No", "Name", "Email", "Phone", "Gender", "Level", "College", "Program", "Program of Studies", "Campus", "STEM / Non-STEM", "Cohort", "Batch", "Scholarship Type", "Category", "Nationality", "Home Country", "District", "Status", "Alumni", "Graduation Year"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0051BA")
    for scholar in scholars:
        ws.append([
            scholar.student_no or "",
            scholar.full_name,
            scholar.email,
            scholar.phone,
            scholar.gender,
            scholar.level_of_study,
            scholar.college,
            scholar.program,
            scholar.program_of_studies,
            scholar.campus,
            scholar.stem_category,
            scholar.cohort,
            scholar.batch,
            scholar.type_of_scholarship,
            scholar.category,
            scholar.nationality,
            scholar.home_country,
            scholar.district,
            scholar.active_status_label,
            "Yes" if scholar.is_alumni else "No",
            scholar.graduation_year or scholar.expected_graduation_year or "",
        ])
    for index, column in enumerate(ws.columns, start=1):
        width = max(len(str(cell.value or "")) for cell in column) + 2
        ws.column_dimensions[get_column_letter(index)].width = min(width, 34)
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="scholars_export.xlsx"'
    wb.save(response)
    return response
