from django.contrib import messages
from django.http import HttpResponse
from django.shortcuts import redirect
from django.shortcuts import render
from django.views.decorators.http import require_http_methods
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.datetime import from_excel
from openpyxl.utils import get_column_letter

from accounts.permissions import DATA_OFFICER, PROGRAM_MANAGER, role_required
from scholars.forms import ScholarForm
from scholars.models import Scholar
from scholars.status import is_active_dropout_status

from .forms import ExcelImportForm


FIELD_NAMES = [
    "student_no", "first_name", "last_name", "gender", "level_of_study", "college", "program",
    "program_of_studies", "campus", "stem_category", "phone", "email", "dob", "marital_status",
    "joining_year", "class_year", "intake", "cohort", "batch", "type_of_scholarship", "category",
    "region_category", "home_country", "nationality", "nationality_country", "province", "district",
    "sector", "cell", "village", "length_of_program", "graduation_year", "home_district", "promotion",
    "accepted_offer_status", "dropout_active_status", "active_status_reason", "replacement",
    "by_enrollment", "alumni_status", "consider_alumni_as_active", "remarks", "date_of_leaving_program",
    "where_is_going", "family_parent_1", "family_parent_2", "family_parent_3",
    "family_parent_relation_1", "family_parent_relation_2", "family_parent_relation_3",
    "family_parent_phone_1", "family_parent_phone_2", "family_parent_phone_3", "life_status",
    "program_type", "is_active", "is_alumni", "expected_graduation_year",
]

HEADER_ALIASES = {
    "student no": "student_no",
    "first name": "first_name",
    "surname": "last_name",
    "last name": "last_name",
    "gender": "gender",
    "level of study": "level_of_study",
    "college": "college",
    "program": "program",
    "program of studies": "program_of_studies",
    "campus": "campus",
    "stem /non stem": "stem_category",
    "stem / non-stem": "stem_category",
    "stem / non stem": "stem_category",
    "stem /non-stem": "stem_category",
    "phone": "phone",
    "for phone": "phone",
    "email": "email",
    "for emails": "email",
    "dob": "dob",
    "marital status": "marital_status",
    "joining year": "joining_year",
    "class": "class_year",
    "intake": "intake",
    "cohort": "cohort",
    "batch": "batch",
    "type of scholarship": "type_of_scholarship",
    "category": "category",
    "region category": "region_category",
    "home country": "home_country",
    "nationality country": "nationality_country",
    "for nationality": "nationality",
    "nationality": "nationality",
    "province": "province",
    "district": "district",
    "sector": "sector",
    "cell": "cell",
    "village": "village",
    "length of program": "length_of_program",
    "graduation year": "graduation_year",
    "home district": "home_district",
    "promotion": "promotion",
    "accepted or rej offer": "accepted_offer_status",
    "accepted or rejected offer": "accepted_offer_status",
    "dropout active": "dropout_active_status",
    "dropout / active status": "dropout_active_status",
    "reason for active/n-active": "active_status_reason",
    "reason for active / non-active": "active_status_reason",
    "replacement": "replacement",
    "by enrolement": "by_enrollment",
    "by enrollment": "by_enrollment",
    "alumni": "alumni_status",
    "by consider alumni as active": "consider_alumni_as_active",
    "consider alumni as active": "consider_alumni_as_active",
    "comments": "remarks",
    "date of leaving program": "date_of_leaving_program",
    "where is going": "where_is_going",
    "family parent 1": "family_parent_1",
    "family arent2": "family_parent_2",
    "family parent 2": "family_parent_2",
    "family parent3": "family_parent_3",
    "family parent 3": "family_parent_3",
    "family parent relation1": "family_parent_relation_1",
    "family parent relation 1": "family_parent_relation_1",
    "family parent relation2": "family_parent_relation_2",
    "family parent relation 2": "family_parent_relation_2",
    "family parent relation3": "family_parent_relation_3",
    "family parent relation 3": "family_parent_relation_3",
    "family parent phone1": "family_parent_phone_1",
    "family parent phone 1": "family_parent_phone_1",
    "family parent phone2": "family_parent_phone_2",
    "family parent phone 2": "family_parent_phone_2",
    "family parent phone3": "family_parent_phone_3",
    "family parent phone 3": "family_parent_phone_3",
    "life status(orphan,parent alive,single parent)": "life_status",
    "life status": "life_status",
}


def parse_bool(value):
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "y", "active", "alumni"}


def clean_phone(value):
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    keep = "".join(ch for ch in text if ch.isdigit() or ch == "+")
    if keep.startswith("0") and len(keep) == 10:
        return "+25" + keep
    if keep.startswith("250"):
        return "+" + keep
    return keep


def parse_year(value):
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        digits = "".join(ch for ch in str(value) if ch.isdigit())
        return int(digits[:4]) if len(digits) >= 4 else None


def parse_date(value):
    if value in (None, ""):
        return None
    if hasattr(value, "date"):
        return value.date()
    try:
        number = float(value)
        if number > 1000:
            return from_excel(number).date()
    except (TypeError, ValueError):
        pass
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            from datetime import datetime
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def canonical_header(header):
    key = " ".join(str(header or "").strip().lower().replace("-", " ").split())
    return HEADER_ALIASES.get(key, key.replace(" ", "_"))


def infer_active(value):
    text = str(value or "").strip().lower()
    if not text:
        return True
    return not any(term in text for term in ["dropout", "inactive", "non-active", "not active"])


def infer_alumni(value):
    text = str(value or "").strip().lower()
    return "alumni" in text and "not yet" not in text


def normalize_row(raw):
    data = {field: raw.get(field) for field in FIELD_NAMES}
    data["phone"] = clean_phone(data.get("phone"))
    for field in ["family_parent_phone_1", "family_parent_phone_2", "family_parent_phone_3"]:
        data[field] = clean_phone(data.get(field))
    data["dob"] = parse_date(data.get("dob"))
    data["date_of_leaving_program"] = parse_date(data.get("date_of_leaving_program"))
    data["consider_alumni_as_active"] = parse_bool(data.get("consider_alumni_as_active"))
    data["joining_year"] = parse_year(data.get("joining_year"))
    data["graduation_year"] = parse_year(data.get("graduation_year"))
    data["expected_graduation_year"] = data["graduation_year"] or parse_year(data.get("expected_graduation_year"))
    if not data.get("program_type"):
        data["program_type"] = "Postgraduate" if str(data.get("level_of_study")).lower().startswith(("master", "phd")) else "Undergraduate"
    if data.get("stem_category"):
        text = str(data["stem_category"]).replace("/", "-").replace(" ", "").upper()
        data["stem_category"] = "Non-STEM" if "NON" in text else "STEM"
    if data.get("nationality") in (None, ""):
        data["nationality"] = data.get("nationality_country") or data.get("home_country") or ""
    data["is_active"] = parse_bool(data.get("is_active")) if data.get("is_active") not in (None, "") else True
    if raw.get("dropout_active_status"):
        data["is_active"] = is_active_dropout_status(raw.get("dropout_active_status"))
    data["is_alumni"] = parse_bool(data.get("is_alumni")) if data.get("is_alumni") not in (None, "") else False
    if raw.get("alumni_status"):
        data["is_alumni"] = infer_alumni(raw.get("alumni_status"))
    return data


def validate_workbook(uploaded):
    wb = load_workbook(uploaded, data_only=True)
    ws = wb.active
    headers = [canonical_header(cell.value) for cell in ws[1]]
    required = ["student_no", "first_name", "last_name", "gender", "level_of_study", "college", "campus", "email"]
    missing = [field for field in required if field not in headers]
    valid_rows = []
    errors = []
    duplicates = 0
    if missing:
        errors.append({"row": 1, "email": "", "message": f"Missing required columns: {', '.join(missing)}"})
        return valid_rows, errors, duplicates
    seen_emails = set()
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        raw = {}
        for header, value in zip(headers, row):
            if header and (header not in raw or raw[header] in (None, "")):
                raw[header] = value
        if not any(raw.values()):
            continue
        data = normalize_row(raw)
        email = str(data.get("email") or "").strip().lower()
        student_no = str(data.get("student_no") or "").strip()
        duplicate_in_file = email in seen_emails or (student_no and student_no in seen_emails)
        duplicate_in_db = Scholar.objects.filter(email=email).exists() or (student_no and Scholar.objects.filter(student_no=student_no).exists())
        if duplicate_in_file or duplicate_in_db:
            duplicates += 1
            errors.append({"row": row_number, "email": data.get("email") or "", "message": "Duplicate Student No or Email already exists in this file or database."})
            continue
        seen_emails.add(email)
        if student_no:
            seen_emails.add(student_no)
        scholar_form = ScholarForm(data)
        if scholar_form.is_valid():
            valid_rows.append({"row": row_number, "data": scholar_form.cleaned_data})
        else:
            errors.append({"row": row_number, "email": data.get("email") or "", "message": "; ".join([f"{field}: {', '.join(errs)}" for field, errs in scholar_form.errors.items()])})
    return valid_rows, errors, duplicates


@role_required(PROGRAM_MANAGER, DATA_OFFICER)
@require_http_methods(["GET", "POST"])
def upload_excel(request):
    form = ExcelImportForm(request.POST or None, request.FILES or None)
    results = None
    preview = request.session.get("import_preview")
    if request.method == "POST" and request.POST.get("action") == "confirm":
        created = 0
        skipped = 0
        errors = []
        for item in preview or []:
            data = item["data"]
            if Scholar.objects.filter(email=data.get("email")).exists() or (data.get("student_no") and Scholar.objects.filter(student_no=data.get("student_no")).exists()):
                skipped += 1
                errors.append({"row": item["row"], "email": data.get("email"), "message": "Duplicate Student No or Email found before save."})
                continue
            scholar_form = ScholarForm(data)
            if scholar_form.is_valid():
                scholar_form.save()
                created += 1
            else:
                errors.append({"row": item["row"], "email": data.get("email"), "message": scholar_form.errors.as_text()})
        request.session.pop("import_preview", None)
        results = {"created": created, "skipped": skipped, "errors": errors}
        if created:
            messages.success(request, f"Imported {created} scholars successfully.")
        if errors:
            messages.warning(request, "Import completed with validation errors.")
    elif request.method == "POST" and form.is_valid():
        valid_rows, errors, duplicates = validate_workbook(form.cleaned_data["file"])
        request.session["import_preview"] = []
        for item in valid_rows:
            serialized = {}
            for key, value in item["data"].items():
                serialized[key] = value.isoformat() if hasattr(value, "isoformat") else (value if value is not None else "")
            request.session["import_preview"].append({"row": item["row"], "data": serialized})
        preview = request.session["import_preview"]
        results = {"created": 0, "skipped": duplicates, "errors": errors, "preview_only": True}
        if preview:
            messages.info(request, f"Preview ready: {len(preview)} valid rows. Confirm to save them.")
    return render(request, "imports/upload.html", {"form": form, "results": results, "fields": FIELD_NAMES, "preview": preview})


@role_required(PROGRAM_MANAGER, DATA_OFFICER)
def sample_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Scholar Import"
    ws.append(FIELD_NAMES)
    ws.append(["Jane", "Doe", "Female", "STEM", "2026", "Bachelor", "Main Campus", "Engineering", "Rwandan", "Undergraduate", True, False, 2030, "Year 1", "May", "+250 700 000 000", "jane.doe@example.com", "Sample row"])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0051BA")
    for index, column in enumerate(ws.columns, start=1):
        width = max(len(str(cell.value or "")) for cell in column) + 2
        ws.column_dimensions[get_column_letter(index)].width = min(width, 32)
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="mcfsp_scholar_import_template.xlsx"'
    wb.save(response)
    return response
