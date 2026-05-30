from io import BytesIO

from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q
from django.http import HttpResponse
from django.shortcuts import render
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from scholars.forms import ScholarFilterForm, filter_scholars
from scholars.models import Scholar
from scholars.status import active_replacement_q, inactive_replacement_q


REPORTS = {
    "scholars": ("Scholar Report", None),
    "campus": ("Campus Report", "campus"),
    "college": ("College Report", "college"),
    "cohort-retention": ("Cohort Retention Report", "cohort"),
    "gender": ("Gender Report", "gender"),
    "nationality": ("Nationality Report", "nationality"),
    "alumni": ("Alumni Report", "is_alumni"),
    "dropout": ("Dropout Report", "inactive"),
}


def report_queryset(request, slug):
    qs, form = filter_scholars(Scholar.objects.all(), request.GET)
    if slug == "alumni":
        qs = qs.filter(is_alumni=True)
    elif slug == "dropout":
        qs = qs.filter(inactive_replacement_q())
    return qs, form


def summary_rows(qs, group_field):
    if not group_field:
        return []
    if group_field == "inactive":
        group_field = "cohort"
    rows = qs.values(group_field).annotate(
        total=Count("id"),
        active=Count("id", filter=active_replacement_q()),
        alumni=Count("id", filter=Q(is_alumni=True)),
        female=Count("id", filter=Q(gender="Female")),
        male=Count("id", filter=Q(gender="Male")),
    ).order_by(group_field)
    return [[row[group_field], row["total"], row["active"], row["alumni"], row["female"], row["male"]] for row in rows]


@login_required
def report_index(request):
    return render(request, "reports/index.html", {"reports": REPORTS})


@login_required
def report_detail(request, slug):
    title, group_field = REPORTS.get(slug, REPORTS["scholars"])
    qs, form = report_queryset(request, slug)
    export = request.GET.get("export")
    if export == "excel":
        return export_excel(title, qs, group_field)
    if export == "pdf":
        return export_pdf(title, qs, group_field)
    total = qs.count()
    active = qs.filter(active_replacement_q()).count()
    inactive = qs.filter(inactive_replacement_q()).count()
    alumni = qs.filter(is_alumni=True).count()
    female = qs.filter(gender="Female").count()
    summary_cards = [
        ("Total Records", total),
        ("Active", active),
        ("Inactive", inactive),
        ("Alumni", alumni),
        ("Female", female),
    ]
    return render(request, "reports/detail.html", {"title": title, "slug": slug, "filter_form": form, "scholars": qs[:500], "summary": summary_rows(qs, group_field), "summary_cards": summary_cards})


def scholar_rows(qs):
    return [[
        s.student_no or "",
        s.full_name,
        s.gender,
        s.level_of_study,
        s.college,
        s.program,
        s.program_of_studies,
        s.campus,
        s.stem_category,
        s.cohort,
        s.batch,
        s.type_of_scholarship,
        s.category,
        s.nationality,
        s.graduation_year or s.expected_graduation_year or "",
        s.active_status_label,
        "Yes" if s.is_alumni else "No",
    ] for s in qs]


def export_excel(title, qs, group_field):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.append([title])
    ws.append([])
    ws["A1"].font = Font(bold=True, size=16, color="0051BA")
    if group_field:
        ws.append(["Group", "Total", "Active", "Alumni", "Female", "Male"])
        for row in summary_rows(qs, group_field):
            ws.append(row)
    else:
        ws.append(["Student No", "Name", "Gender", "Level", "College", "Program", "Program of Studies", "Campus", "STEM / Non-STEM", "Cohort", "Batch", "Scholarship Type", "Category", "Nationality", "Graduation Year", "Status", "Alumni"])
        for row in scholar_rows(qs):
            ws.append(row)
    header_row = 3
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0051BA")
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = ws.dimensions
    for index, column in enumerate(ws.columns, start=1):
        width = max(len(str(cell.value or "")) for cell in column) + 2
        ws.column_dimensions[get_column_letter(index)].width = min(width, 34)
    output = BytesIO()
    wb.save(output)
    response = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{title.lower().replace(" ", "_")}.xlsx"'
    return response


def export_pdf(title, qs, group_field):
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=28, bottomMargin=24, leftMargin=28, rightMargin=28)
    styles = getSampleStyleSheet()
    data = [["Group", "Total", "Active", "Alumni", "Female", "Male"]] + summary_rows(qs, group_field) if group_field else [["Student No", "Name", "Gender", "Level", "College", "Program", "Campus", "STEM", "Cohort", "Scholarship", "Category", "Nationality", "Grad Year", "Status", "Alumni"]] + [row[:6] + [row[7], row[8], row[9], row[11], row[12], row[13], row[14], row[15], row[16]] for row in scholar_rows(qs[:200])]
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0051BA")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 8),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7f5f1")]),
        ("FONT", (0, 0), (-1, -1), "Helvetica", 8),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story = [
        Paragraph(f"<b>{title}</b>", styles["Title"]),
        Paragraph(f"Generated from live Scholar records. Total records: {qs.count()}", styles["Normal"]),
        Spacer(1, 14),
        table,
    ]
    doc.build(story)
    response = HttpResponse(output.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{title.lower().replace(" ", "_")}.pdf"'
    return response
